from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import datetime
import sys
import os

sys.path.insert(0, os.path.dirname(__file__))
from styles import (
    mk_font, mk_fill, thin_border, medium_bottom_border,
    ALIGN_CENTER, ALIGN_RIGHT, ALIGN_LEFT,
    NUMBER_FMT, DATE_FMT,
    C_HEADER_BG, C_HEADER_FG, C_VENDOR_BG, C_VENDOR_FG,
    C_OPEN_BG, C_CLOSE_BG, C_PURCHASE_BG, C_PAYMENT_BG,
    C_JOURNAL_BG, C_DEBIT_BG, C_RB_POS, C_RB_NEG, C_RB_ZERO
)

COLS       = ["Date", "Vch Type", "Vch No.", "Particulars / Narration",
              "Debit (₹)", "Credit (₹)", "Running Balance (₹)"]
COL_WIDTHS = [14, 28, 20, 48, 14, 14, 20]


def write_output_xlsx(vendor_data: dict, output_path: str,
                      ledger_type: str = "creditor",
                      include_summary: bool = True,
                      include_outstanding: bool = True,
                      progress_callback=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger with Running Balance"

    _write_headers(ws, ledger_type)
    current_row = 4

    vendors = list(vendor_data.items())
    for idx, (vname, data) in enumerate(vendors):
        if progress_callback:
            progress_callback(idx + 1, len(vendors), vname)
        current_row = _write_vendor_block(ws, current_row, vname, data, ledger_type)
        current_row += 1

    if include_summary:
        _write_summary_sheet(wb, vendor_data, ledger_type)

    if include_outstanding:
        _write_outstanding_sheet(wb, vendor_data, ledger_type)

    wb.save(output_path)


def _rb_color(rb):
    if rb > 0.5:   return C_RB_POS
    if rb < -0.5:  return C_RB_NEG
    return C_RB_ZERO


def _tx_bg(vt, ledger_type):
    if ledger_type == "debtor":
        if "Receipt" in vt or "Payment" in vt: return C_PAYMENT_BG
        if "Journal" in vt:                     return C_JOURNAL_BG
        if "Debit Note" in vt or "Credit Note" in vt: return C_DEBIT_BG
        return C_PURCHASE_BG
    else:
        if "Purchase"   in vt: return C_PURCHASE_BG
        if "Payment"    in vt: return C_PAYMENT_BG
        if "Journal"    in vt: return C_JOURNAL_BG
        if "Debit Note" in vt: return C_DEBIT_BG
        return "FFFFFF"


def _write_vendor_block(ws, start_row, vname, data, ledger_type):
    r        = start_row
    opening  = data["opening"]
    txns     = data["transactions"]
    balances = data["running_balances"]
    sc       = data.get("stated_closing") or 0
    cc       = data.get("computed_closing", 0)

    ws.merge_cells(f"A{r}:G{r}")
    c = ws.cell(row=r, column=1, value=f"  {vname}")
    c.font = mk_font(bold=True, color=C_VENDOR_FG, size=11)
    c.fill = mk_fill(C_VENDOR_BG)
    c.alignment = ALIGN_LEFT
    c.border = medium_bottom_border()
    ws.row_dimensions[r].height = 18
    r += 1

    _write_balance_row(ws, r, datetime.date(2025, 4, 1), "Opening Balance",
                       opening, C_OPEN_BG, ledger_type=ledger_type)
    r += 1

    for i, (dt, vt, vno, nar, deb, cre) in enumerate(txns):
        rb       = balances[i]
        bg       = _tx_bg(vt, ledger_type)
        rb_col   = _rb_color(rb)
        date_val = dt.date() if hasattr(dt, "date") else dt

        row_vals = [
            (date_val,    mk_font(),                                     bg, ALIGN_CENTER, DATE_FMT),
            (vt,          mk_font(),                                     bg, ALIGN_LEFT,   None),
            (vno,         mk_font(),                                     bg, ALIGN_LEFT,   None),
            (nar,         mk_font(),                                     bg, ALIGN_LEFT,   None),
            (deb or None, mk_font(color="922B21") if deb else mk_font(), bg, ALIGN_RIGHT,  NUMBER_FMT),
            (cre or None, mk_font(color="1A5276") if cre else mk_font(), bg, ALIGN_RIGHT,  NUMBER_FMT),
            (rb,          mk_font(bold=True, color=rb_col),              bg, ALIGN_RIGHT,  NUMBER_FMT),
        ]
        for ci, (val, fnt, bg2, aln, fmt) in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = fnt; cell.fill = mk_fill(bg2)
            cell.alignment = aln; cell.border = thin_border()
            if fmt: cell.number_format = fmt
        r += 1

    match_note = "✓ MATCH" if abs(sc - cc) < 0.5 else f"⚠ OFF by {sc - cc:+,.2f}"
    _write_balance_row(ws, r, datetime.date(2027, 4, 30), "Closing Balance",
                       sc, C_CLOSE_BG, note=match_note, ledger_type=ledger_type)
    r += 1
    return r


def _write_balance_row(ws, r, date_val, label, amount, bg, note="", ledger_type="creditor"):
    rb_col     = _rb_color(amount)
    note_color = "1E8449" if "MATCH" in note else ("922B21" if "OFF" in note else "000000")

    # For debtors: positive balance sits in debit col (receivable)
    # For creditors: positive balance sits in credit col (payable)
    if ledger_type == "debtor":
        debit_val  = amount if amount >= 0 else None
        credit_val = None
    else:
        debit_val  = None
        credit_val = amount if amount >= 0 else None

    row_data = [
        (date_val,   mk_font(bold=True),                   ALIGN_CENTER, DATE_FMT),
        (label,      mk_font(bold=True),                   ALIGN_LEFT,   None),
        ("",         mk_font(),                            ALIGN_LEFT,   None),
        (note,       mk_font(bold=True, color=note_color), ALIGN_LEFT,   None),
        (debit_val,  mk_font(bold=True),                   ALIGN_RIGHT,  NUMBER_FMT),
        (credit_val, mk_font(bold=True),                   ALIGN_RIGHT,  NUMBER_FMT),
        (amount,     mk_font(bold=True, color=rb_col),     ALIGN_RIGHT,  NUMBER_FMT),
    ]
    for ci, (val, fnt, aln, fmt) in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=ci, value=val)
        cell.font = fnt; cell.fill = mk_fill(bg)
        cell.alignment = aln; cell.border = thin_border()
        if fmt: cell.number_format = fmt


def _write_headers(ws, ledger_type="creditor"):
    if ledger_type == "debtor":
        title    = "SUNDRY DEBTORS STATEMENT — WITH RUNNING BALANCE"
        subtitle = ("Period: 1-Apr-2025 to 30-Apr-2027  |  "
                    "Running Balance: positive = customer OWES you; negative = you owe customer (excess receipt)")
    else:
        title    = "SUNDRY CREDITORS STATEMENT — WITH RUNNING BALANCE"
        subtitle = ("Period: 1-Apr-2025 to 30-Apr-2027  |  "
                    "Running Balance: positive = you OWE vendor; negative = vendor OWES you")

    ws.merge_cells("A1:G1")
    ws["A1"] = title
    ws["A1"].font      = mk_font(bold=True, color=C_HEADER_FG, size=13)
    ws["A1"].fill      = mk_fill(C_HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    ws["A2"] = subtitle
    ws["A2"].font      = mk_font(italic=True, color="444444", size=9)
    ws["A2"].fill      = mk_fill("EBF3FB")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    for ci, (h, w) in enumerate(zip(COLS, COL_WIDTHS), 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font      = mk_font(bold=True, color=C_HEADER_FG, size=10)
        cell.fill      = mk_fill(C_HEADER_BG)
        cell.alignment = ALIGN_CENTER
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A4"


def _write_summary_sheet(wb, vendor_data, ledger_type="creditor"):
    ws = wb.create_sheet("Summary")

    title      = "SUNDRY DEBTORS — SUMMARY" if ledger_type == "debtor" else "SUNDRY CREDITORS — SUMMARY"
    bal_header = "Closing Receivable (₹)"   if ledger_type == "debtor" else "Closing Balance (₹)"
    party_hdr  = "Party Name"               if ledger_type == "debtor" else "Vendor Name"
    total_lbl  = "TOTAL RECEIVABLE"         if ledger_type == "debtor" else "TOTAL OUTSTANDING"

    ws.merge_cells("A1:E1")
    ws["A1"] = title
    ws["A1"].font      = mk_font(bold=True, color=C_HEADER_FG, size=13)
    ws["A1"].fill      = mk_fill(C_HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["#", party_hdr, "Opening Balance (₹)", bal_header, "Status"]
    widths  = [5, 45, 22, 22, 18]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font      = mk_font(bold=True, color=C_HEADER_FG, size=10)
        cell.fill      = mk_fill(C_HEADER_BG)
        cell.alignment = ALIGN_CENTER
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 18

    total_closing = 0.0
    for idx, (vname, data) in enumerate(vendor_data.items(), 1):
        r  = idx + 2
        ob = data["opening"]
        cc = data.get("computed_closing", 0)
        total_closing += cc

        if cc > 0.5:
            status     = "Receivable" if ledger_type == "debtor" else "Payable"
            status_col = C_RB_POS
            row_bg     = "FFFFFF" if idx % 2 == 0 else "F5F5F5"
        elif cc < -0.5:
            status     = "Credit Bal ⚠" if ledger_type == "debtor" else "Overpaid ⚠"
            status_col = C_RB_NEG
            row_bg     = "FFF0F0"
        else:
            status     = "Settled ✓"
            status_col = C_RB_ZERO
            row_bg     = "F0FFF0"

        for ci, val in enumerate([idx, vname, ob, cc, status], 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.fill = mk_fill(row_bg); cell.border = thin_border()
            if ci == 5:
                cell.font = mk_font(bold=True, color=status_col); cell.alignment = ALIGN_CENTER
            elif ci in (3, 4):
                cell.font = mk_font(color=_rb_color(cc) if ci == 4 else "000000")
                cell.alignment = ALIGN_RIGHT; cell.number_format = NUMBER_FMT
            elif ci == 1:
                cell.font = mk_font(color="666666"); cell.alignment = ALIGN_CENTER
            else:
                cell.font = mk_font(); cell.alignment = ALIGN_LEFT

    total_r = len(vendor_data) + 3
    ws.merge_cells(f"A{total_r}:B{total_r}")
    ws.cell(row=total_r, column=1, value=total_lbl).font = mk_font(bold=True, color=C_HEADER_FG, size=10)
    ws.cell(row=total_r, column=1).fill      = mk_fill(C_HEADER_BG)
    ws.cell(row=total_r, column=1).alignment = ALIGN_RIGHT
    for ci in (2, 3, 4, 5):
        ws.cell(row=total_r, column=ci).fill = mk_fill(C_HEADER_BG)
    total_cell = ws.cell(row=total_r, column=4, value=round(total_closing, 2))
    total_cell.font = mk_font(bold=True, color=C_HEADER_FG, size=11)
    total_cell.alignment = ALIGN_RIGHT; total_cell.number_format = NUMBER_FMT

    ws.freeze_panes = "A3"



# ── Outstanding Breakdown sheet ──────────────────────────────────────────────

_OB_REASON_FILLS = {
    "NEVER_TOUCHED":            "FDECEA",
    "PARTIALLY_PAID":           "FFF8E7",
    "POST_LAST_PAYMENT":        "E3F2FD",
    "RATE_REBATE_TAIL":         "F1F8E9",
    "OPENING_BALANCE_RESIDUAL": "F3E5F5",
}

_OB_COLS   = ["#", "Date", "Vch Type", "Vch No.", "Original (₹)",
              "Paid So Far (₹)", "Remaining (₹)", "Reason"]
_OB_WIDTHS = [5, 14, 30, 22, 16, 16, 16, 70]


def _ob_vendor_header(ws, r, vname):
    ws.merge_cells(f"A{r}:H{r}")
    c = ws.cell(row=r, column=1, value=f"  {vname.strip()}")
    c.font      = mk_font(bold=True, color="FFFFFF", size=11)
    c.fill      = mk_fill("1F3864")
    c.alignment = ALIGN_LEFT
    c.border    = medium_bottom_border()
    ws.row_dimensions[r].height = 18
    return r + 1


def _ob_summary_row(ws, r, text, bg="DDEEFF", height=30):
    ws.merge_cells(f"A{r}:H{r}")
    c = ws.cell(row=r, column=1, value=text)
    c.font      = mk_font(italic=True, color="1A3A5C", size=10)
    c.fill      = mk_fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = height
    return r + 1


def _ob_col_headers(ws, r):
    for ci, h in enumerate(_OB_COLS, 1):
        cell = ws.cell(row=r, column=ci, value=h)
        cell.font      = mk_font(bold=True, color=C_HEADER_FG, size=10)
        cell.fill      = mk_fill(C_HEADER_BG)
        cell.alignment = ALIGN_CENTER
        cell.border    = thin_border()
    ws.row_dimensions[r].height = 18
    return r + 1


def _ob_note_row(ws, r, text, bg="FFF8E7"):
    ws.merge_cells(f"A{r}:H{r}")
    c = ws.cell(row=r, column=1, value=text)
    c.font      = mk_font(italic=True, color="922B21", size=9)
    c.fill      = mk_fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 24
    return r + 1


def _ob_item_row(ws, r, seq, item):
    remaining = item["original"] - item["consumed"]
    rtype     = item.get("reason_type") or "NEVER_TOUCHED"
    bg        = _OB_REASON_FILLS.get(rtype, "FFFFFF")
    d         = item["date"]
    date_val  = d.date() if hasattr(d, "date") else d

    row_vals = [
        (seq,              mk_font(color="666666"), ALIGN_CENTER, None),
        (date_val,         mk_font(),               ALIGN_CENTER, DATE_FMT),
        (item["vch_type"], mk_font(),               ALIGN_LEFT,   None),
        (item["vch_no"] if item["vch_no"] != "Opening Balance" else "",
                           mk_font(),               ALIGN_LEFT,   None),
        (item["original"], mk_font(),               ALIGN_RIGHT,  NUMBER_FMT),
        (item["consumed"], mk_font(),               ALIGN_RIGHT,  NUMBER_FMT),
        (round(remaining, 2),
                           mk_font(bold=True),      ALIGN_RIGHT,  NUMBER_FMT),
        (item.get("reason_text", ""),
                           mk_font(size=9),
                           Alignment(horizontal="left", vertical="center", wrap_text=True),
                           None),
    ]
    for ci, (val, fnt, aln, fmt) in enumerate(row_vals, 1):
        cell = ws.cell(row=r, column=ci, value=val)
        cell.font      = fnt
        cell.fill      = mk_fill(bg)
        cell.alignment = aln
        cell.border    = thin_border()
        if fmt:
            cell.number_format = fmt
    ws.row_dimensions[r].height = 36
    return r + 1


def _write_outstanding_sheet(wb, vendor_data: dict, ledger_type: str = "creditor"):
    ws = wb.create_sheet("Outstanding Breakdown")

    title = ("SUNDRY DEBTORS — OUTSTANDING BREAKDOWN"
             if ledger_type == "debtor"
             else "SUNDRY CREDITORS — OUTSTANDING BREAKDOWN")

    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font      = mk_font(bold=True, color=C_HEADER_FG, size=13)
    ws["A1"].fill      = mk_fill(C_HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for ci, w in enumerate(_OB_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    r = 2

    for vname, data in vendor_data.items():
        fifo_items = data.get("fifo_items")
        if fifo_items is None:
            continue  # FIFO not run for this vendor

        pattern = data.get("fifo_pattern", "")
        cc      = data.get("computed_closing", 0)
        summary = data.get("vendor_summary_sentence", "")

        # SKIP: zero/negative closing — don't show in sheet at all
        if pattern == "SKIP":
            continue

        # Vendor header
        r = _ob_vendor_header(ws, r, vname)

        # Summary sentence
        r = _ob_summary_row(ws, r, summary)

        # Column headers
        r = _ob_col_headers(ws, r)

        # Item rows — only items with remaining > 0.01
        seq = 0
        for item in fifo_items:
            remaining = item["original"] - item["consumed"]
            if remaining <= 0.01:
                continue
            seq += 1
            r = _ob_item_row(ws, r, seq, item)

        r += 1  # spacer between vendors

    ws.freeze_panes = "A2"
