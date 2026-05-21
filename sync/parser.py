import io
from openpyxl import load_workbook


def _num(v):
    return float(v) if v and isinstance(v, (int, float)) else 0.0


def _open_workbook(filepath: str):
    """
    Open a workbook regardless of extension.
    Tally exports raw .xls files that are actually Office Open XML (zip/xlsx)
    — openpyxl rejects them based on the .xls extension alone.
    We detect this by reading the magic bytes: PK (zip) = xlsx-in-disguise,
    D0CF = genuine BIFF8 .xls (not supported by openpyxl; user must convert).
    """
    with open(filepath, "rb") as f:
        magic = f.read(4)
        f.seek(0)
        raw = f.read()

    if magic[:2] == b"PK":
        # ZIP-based file → load directly via BytesIO (bypasses extension check)
        return load_workbook(io.BytesIO(raw), read_only=True)
    else:
        # Try normally (handles .xlsx / .xlsm with correct extension)
        return load_workbook(filepath, read_only=True)


def detect_ledger_type(rows) -> str:
    """
    Auto-detect whether this is a creditors or debtors export
    by reading the report title row.
    Returns 'debtor' or 'creditor'.
    """
    for row in rows[:5]:
        vals = [v for v in row if v is not None]
        if vals and isinstance(vals[0], str):
            title = vals[0].lower()
            if "debtor" in title:
                return "debtor"
    return "creditor"


def parse_workbook(filepath: str) -> tuple:
    """
    Returns (vendor_dict, ledger_type).
    vendor_dict: {name: {opening, stated_closing, transactions}}
    transactions: list of (date, vch_type, vch_no, narration, debit, credit)

    For CREDITORS:
      opening/closing: col[3] - col[2]  (credit-normal: we owe them)
      debit col reduces liability, credit col increases it

    For DEBTORS:
      opening/closing: col[2] - col[3]  (debit-normal: they owe us)
      debit col increases receivable, credit col reduces it
    """
    wb = _open_workbook(filepath)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    ledger_type = detect_ledger_type(rows)

    vendors = {}
    current = None

    SKIP = {"Sundry Creditors Statement", "Sundry Debtors Statement", "Date"}

    for row in rows:
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        # Party name detection
        if (len(vals) == 1
                and isinstance(vals[0], str)
                and vals[0] not in SKIP
                and not str(vals[0]).startswith("Period")
                and not str(vals[0]).startswith("--")):
            current = vals[0]
            vendors[current] = {
                "opening": 0.0,
                "stated_closing": None,
                "transactions": []
            }
            continue

        if not current:
            continue

        vch_type = str(row[1]) if row[1] else ""

        if "Opening Balance" in vch_type:
            col2, col3 = _num(row[2]), _num(row[3])
            if col2 == 0.0 and col3 == 0.0:
                # Tally stores negative opening only in col G (running balance)
                col_g = _num(row[6]) if len(row) > 6 and row[6] is not None else 0.0
                vendors[current]["opening"] = col_g
            elif ledger_type == "debtor":
                vendors[current]["opening"] = col2 - col3   # debit-normal
            else:
                vendors[current]["opening"] = col3 - col2   # credit-normal

        elif "Closing Balance" in vch_type:
            col2, col3 = _num(row[2]), _num(row[3])
            if ledger_type == "debtor":
                vendors[current]["stated_closing"] = col2 - col3
            else:
                vendors[current]["stated_closing"] = col3 - col2

        else:
            dt  = row[0]
            vno = row[2] if row[2] else ""
            nar = row[3] if row[3] else ""
            deb = _num(row[4])
            cre = _num(row[5])
            vendors[current]["transactions"].append(
                (dt, vch_type, str(vno), str(nar), deb, cre)
            )

    wb.close()
    return vendors, ledger_type
