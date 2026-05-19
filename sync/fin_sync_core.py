"""
fin_sync_core.py — PriceMaster Financial Sync core logic.
No GUI code here. Imported by fin_sync_tool_tab.py and sync_tool.pyw.

Public API:
    sync_debtors(excel_path, supabase_url, service_role_key, log_callback)
        -> (bool, str)
    sync_creditors(excel_path, supabase_url, service_role_key, log_callback)
        -> (bool, str)
    sync_address_book(excel_path, supabase_url, service_role_key, log_callback)
        -> (bool, str)

Parsing rules (verified against live Tally exports):
  - Party name rows: the name cell has leading whitespace; ALL other columns
    in that row are None. Detection: strip all cells, check rest are all None.
    Do NOT rely on exact leading-space count — it varies.
  - Summary row patterns to skip: row[0] is a string starting with "Total",
    "Note:", "Period:", or the sheet title. Also skip if row[0] is None.
  - Outstanding breakdown has interleaved header rows (#, Date, Vch Type, …)
    between parties — detect by row[0] == '#' or isinstance(row[0], int).
  - Dates arrive as datetime.datetime objects from openpyxl — convert to
    ISO string 'YYYY-MM-DD'. None dates are stored as None.
  - Amounts: None → 0. Negative values are kept as-is (they carry meaning).
  - Chunk size: 400 rows per Supabase request (matches existing sync_core.py).
"""

import time
from datetime import datetime, date

from openpyxl import load_workbook
from supabase import create_client


# ── Constants ─────────────────────────────────────────────────────────────────

CHUNK_SIZE = 400


# ── Helpers ───────────────────────────────────────────────────────────────────

def _log(callback, message: str):
    if callback:
        callback(message)


def _admin_client(supabase_url: str, service_role_key: str):
    return create_client(supabase_url, service_role_key)


def _to_date(val) -> str | None:
    """Convert openpyxl datetime/date or None to ISO string or None."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d")
        return val.isoformat()
    # Occasionally comes in as a string already
    if isinstance(val, str) and val.strip():
        return val.strip()
    return None


def _to_num(val) -> float:
    """Coerce None/empty to 0.0, else float."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _is_party_name_row(row: tuple) -> bool:
    """
    A party-name row has the party name (possibly with leading whitespace)
    in position 0, and ALL other columns are None.
    Guard: the name must be a non-empty string after stripping.
    """
    if not row:
        return False
    first = row[0]
    if not isinstance(first, str) or not first.strip():
        return False
    rest = row[1:]
    return all(v is None for v in rest)


def _is_all_none(row: tuple) -> bool:
    return all(v is None for v in row)


def _chunk(lst: list, size: int):
    for i in range(0, len(lst), size):
        yield lst[i : i + size]


def _insert_chunks(sb, table: str, rows: list, log_callback=None):
    """Insert rows into table in CHUNK_SIZE batches. Returns total inserted."""
    total = 0
    for batch in _chunk(rows, CHUNK_SIZE):
        sb.table(table).insert(batch).execute()
        total += len(batch)
        _log(log_callback, f"  ↳ Inserted {total}/{len(rows)} rows into {table}…")
    return total


# ── Ledger parser (shared by debtors and creditors) ───────────────────────────

def _parse_ledger_sheet(ws, party_type: str, log_callback=None) -> list[dict]:
    """
    Parse 'Ledger with Running Balance' sheet.
    Returns list of dicts ready for fin_ledger insert.

    Sheet structure:
      Row 1: title string
      Row 2: period string
      Row 3: column headers
      Row 4+: party-name rows, then transaction rows, then blank rows
    """
    rows_out = []
    current_party = None
    skipped_header_rows = 0

    for raw_row in ws.iter_rows(min_row=4, values_only=True):
        # Pad to 7 to avoid index errors
        row = tuple(raw_row) + (None,) * 7
        row = row[:7]

        if _is_all_none(row):
            continue

        if _is_party_name_row(row):
            current_party = row[0].strip()
            continue

        # Skip column header rows (they repeat between parties in some exports)
        if isinstance(row[0], str) and row[0].strip() in ("Date", "#"):
            skipped_header_rows += 1
            continue

        # Skip summary/note rows (strings in col 0 that aren't dates or amounts)
        if isinstance(row[0], str):
            stripped = row[0].strip()
            if (stripped.startswith(("Total", "Note:", "Period:", "SUNDRY"))
                    or stripped == ""):
                continue
            # If it looks like a date-string, continue parsing (handled below)
            # Otherwise skip unknown text rows
            if not any(c.isdigit() for c in stripped):
                continue

        if current_party is None:
            continue

        txn_date = _to_date(row[0])
        vch_type = str(row[1]).strip() if row[1] is not None else None
        vch_no = str(row[2]).strip() if row[2] is not None else None
        narration = str(row[3]).strip() if row[3] is not None else None
        debit = _to_num(row[4])
        credit = _to_num(row[5])
        balance = _to_num(row[6])

        rows_out.append({
            "party_type": party_type,
            "party_name": current_party,
            "txn_date": txn_date,
            "vch_type": vch_type,
            "vch_no": vch_no,
            "narration": narration,
            "debit": debit,
            "credit": credit,
            "balance": balance,
        })

    if skipped_header_rows:
        _log(log_callback, f"  (Skipped {skipped_header_rows} repeated header rows in ledger)")

    return rows_out


# ── Summary parser (shared) ───────────────────────────────────────────────────

def _parse_summary_sheet(ws, party_type: str) -> list[dict]:
    """
    Parse 'Summary' sheet.
    Row 1: title; Row 2: column headers; Row 3+: party data.
    Column order: #, Party Name / Vendor Name, Opening Balance (₹),
                  Closing Receivable/Balance (₹), Status
    Returns list of dicts for fin_parties.
    """
    parties = []
    for raw_row in ws.iter_rows(min_row=3, values_only=True):
        row = tuple(raw_row) + (None,) * 5
        row = row[:5]

        # Skip blank rows and any stray header-like rows
        if row[1] is None or not isinstance(row[1], str):
            continue
        party_name = row[1].strip()
        if not party_name:
            continue

        opening_bal = _to_num(row[2])
        closing_bal = _to_num(row[3])
        status = str(row[4]).strip() if row[4] is not None else None

        parties.append({
            "party_type": party_type,
            "party_name": party_name,
            "opening_bal": opening_bal,
            "closing_bal": closing_bal,
            "status": status,
        })

    return parties


# ── Outstanding parser (shared) ───────────────────────────────────────────────

def _parse_outstanding_sheet(ws, party_type: str, log_callback=None) -> list[dict]:
    """
    Parse 'Outstanding Breakdown' sheet.
    Structure: title row → party-name row → summary-text row →
               header row (#, Date, …) → invoice rows → blank → repeat.
    Returns list of dicts for fin_outstanding.
    """
    rows_out = []
    current_party = None

    for raw_row in ws.iter_rows(min_row=2, values_only=True):
        row = tuple(raw_row) + (None,) * 8
        row = row[:8]

        if _is_all_none(row):
            continue

        if _is_party_name_row(row):
            current_party = row[0].strip()
            continue

        # Skip header rows and summary/note text rows
        if isinstance(row[0], str):
            stripped = row[0].strip()
            if stripped in ("#", "Date") or stripped.startswith(
                ("Total", "Note:", "SUNDRY", "Period:")
            ):
                continue
            if not any(c.isdigit() for c in stripped):
                continue

        # Skip repeated column-header rows where row[0] == '#'
        if row[0] == "#":
            continue

        # Invoice rows: row[0] is an integer (the # column)
        if not isinstance(row[0], int):
            continue

        if current_party is None:
            continue

        inv_date = _to_date(row[1])
        vch_type = str(row[2]).strip() if row[2] is not None else None
        vch_no = str(row[3]).strip() if row[3] is not None else None
        original_amt = _to_num(row[4])
        paid_amt = _to_num(row[5])
        remaining = _to_num(row[6])
        reason = str(row[7]).strip() if row[7] is not None else None

        rows_out.append({
            "party_type": party_type,
            "party_name": current_party,
            "inv_date": inv_date,
            "vch_type": vch_type,
            "vch_no": vch_no,
            "original_amt": original_amt,
            "paid_amt": paid_amt,
            "remaining": remaining,
            "reason": reason,
        })

    return rows_out


# ── Cross-validation helper ────────────────────────────────────────────────────

def _cross_validate(summary_parties: list[dict], ledger_rows: list[dict],
                    log_callback=None):
    """
    Warn if a party in summary has non-zero closing balance but zero ledger rows.
    This catches silent parse failures.
    """
    party_in_ledger = set(r["party_name"].lower() for r in ledger_rows)
    warnings = 0
    for p in summary_parties:
        if p["closing_bal"] != 0 and p["party_name"].lower() not in party_in_ledger:
            _log(log_callback,
                 f"  ⚠ WARNING: '{p['party_name']}' has closing bal "
                 f"{p['closing_bal']} but ZERO ledger rows found — check Excel.")
            warnings += 1
    if warnings == 0:
        _log(log_callback, "  ✓ Cross-validation passed — all non-zero parties have ledger rows.")
    else:
        _log(log_callback, f"  ⚠ {warnings} party/ies with closing balance but no ledger rows.")


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  Public API                                                                 ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

def sync_debtors(
    excel_path: str,
    supabase_url: str,
    service_role_key: str,
    log_callback=None,
) -> tuple[bool, str]:
    """
    Parse Sundry Debtors Excel and push to Supabase.
    Returns (success: bool, summary_message: str).
    """
    t_start = time.time()
    party_type = "debtor"

    try:
        _log(log_callback, "📂 Loading Sundry Debtors Excel…")
        wb = load_workbook(excel_path, read_only=True, data_only=True)

        # ── Parse ─────────────────────────────────────────────────────────
        _log(log_callback, "📊 Parsing Summary sheet…")
        summary_parties = _parse_summary_sheet(wb["Summary"], party_type)
        _log(log_callback, f"  → Found {len(summary_parties)} parties in Summary.")

        _log(log_callback, "📊 Parsing Ledger sheet…")
        ledger_rows = _parse_ledger_sheet(
            wb["Ledger with Running Balance"], party_type, log_callback
        )
        _log(log_callback, f"  → Found {len(ledger_rows)} ledger rows.")

        _log(log_callback, "📊 Parsing Outstanding Breakdown sheet…")
        outstanding_rows = _parse_outstanding_sheet(
            wb["Outstanding Breakdown"], party_type, log_callback
        )
        _log(log_callback, f"  → Found {len(outstanding_rows)} outstanding invoice rows.")

        wb.close()

        # ── Cross-validate ─────────────────────────────────────────────────
        _log(log_callback, "🔍 Cross-validating parties vs ledger…")
        _cross_validate(summary_parties, ledger_rows, log_callback)

        # ── Upload ────────────────────────────────────────────────────────
        _log(log_callback, "🔌 Connecting to Supabase…")
        sb = _admin_client(supabase_url, service_role_key)

        _log(log_callback, f"🗑  Deleting existing debtor data from fin_parties…")
        sb.table("fin_parties").delete().eq("party_type", party_type).execute()

        _log(log_callback, f"🗑  Deleting existing debtor data from fin_ledger…")
        sb.table("fin_ledger").delete().eq("party_type", party_type).execute()

        _log(log_callback, f"🗑  Deleting existing debtor data from fin_outstanding…")
        sb.table("fin_outstanding").delete().eq("party_type", party_type).execute()

        _log(log_callback, f"📤 Inserting {len(summary_parties)} parties…")
        _insert_chunks(sb, "fin_parties", summary_parties, log_callback)

        _log(log_callback, f"📤 Inserting {len(ledger_rows)} ledger rows…")
        _insert_chunks(sb, "fin_ledger", ledger_rows, log_callback)

        if outstanding_rows:
            _log(log_callback, f"📤 Inserting {len(outstanding_rows)} outstanding rows…")
            _insert_chunks(sb, "fin_outstanding", outstanding_rows, log_callback)
        else:
            _log(log_callback, "  (No outstanding rows to insert — all debtors settled.)")

        # ── Sync log ──────────────────────────────────────────────────────
        sb.table("fin_sync_log").insert({
            "file_type": "debtors",
            "row_count": len(ledger_rows),
            "party_count": len(summary_parties),
            "status": "success",
        }).execute()

        elapsed = time.time() - t_start
        msg = (
            f"✅ Debtors synced — {len(summary_parties)} parties, "
            f"{len(ledger_rows)} ledger rows, "
            f"{len(outstanding_rows)} outstanding rows "
            f"({elapsed:.1f}s)"
        )
        _log(log_callback, msg)
        return True, msg

    except Exception as exc:
        msg = f"❌ Debtors sync failed: {exc}"
        _log(log_callback, msg)
        try:
            _admin_client(supabase_url, service_role_key).table("fin_sync_log").insert({
                "file_type": "debtors",
                "row_count": 0,
                "party_count": 0,
                "status": f"error: {exc}",
            }).execute()
        except Exception:
            pass
        return False, msg


def sync_creditors(
    excel_path: str,
    supabase_url: str,
    service_role_key: str,
    log_callback=None,
) -> tuple[bool, str]:
    """
    Parse Sundry Creditors Excel and push to Supabase.
    Returns (success: bool, summary_message: str).
    """
    t_start = time.time()
    party_type = "creditor"

    try:
        _log(log_callback, "📂 Loading Sundry Creditors Excel…")
        wb = load_workbook(excel_path, read_only=True, data_only=True)

        _log(log_callback, "📊 Parsing Summary sheet…")
        summary_parties = _parse_summary_sheet(wb["Summary"], party_type)
        _log(log_callback, f"  → Found {len(summary_parties)} parties in Summary.")

        _log(log_callback, "📊 Parsing Ledger sheet…")
        ledger_rows = _parse_ledger_sheet(
            wb["Ledger with Running Balance"], party_type, log_callback
        )
        _log(log_callback, f"  → Found {len(ledger_rows)} ledger rows.")

        _log(log_callback, "📊 Parsing Outstanding Breakdown sheet…")
        outstanding_rows = _parse_outstanding_sheet(
            wb["Outstanding Breakdown"], party_type, log_callback
        )
        _log(log_callback, f"  → Found {len(outstanding_rows)} outstanding invoice rows.")

        wb.close()

        _log(log_callback, "🔍 Cross-validating parties vs ledger…")
        _cross_validate(summary_parties, ledger_rows, log_callback)

        _log(log_callback, "🔌 Connecting to Supabase…")
        sb = _admin_client(supabase_url, service_role_key)

        _log(log_callback, "🗑  Deleting existing creditor data from fin_parties…")
        sb.table("fin_parties").delete().eq("party_type", party_type).execute()

        _log(log_callback, "🗑  Deleting existing creditor data from fin_ledger…")
        sb.table("fin_ledger").delete().eq("party_type", party_type).execute()

        _log(log_callback, "🗑  Deleting existing creditor data from fin_outstanding…")
        sb.table("fin_outstanding").delete().eq("party_type", party_type).execute()

        _log(log_callback, f"📤 Inserting {len(summary_parties)} parties…")
        _insert_chunks(sb, "fin_parties", summary_parties, log_callback)

        _log(log_callback, f"📤 Inserting {len(ledger_rows)} ledger rows…")
        _insert_chunks(sb, "fin_ledger", ledger_rows, log_callback)

        if outstanding_rows:
            _log(log_callback, f"📤 Inserting {len(outstanding_rows)} outstanding rows…")
            _insert_chunks(sb, "fin_outstanding", outstanding_rows, log_callback)
        else:
            _log(log_callback, "  (No outstanding rows to insert — all creditors settled.)")

        sb.table("fin_sync_log").insert({
            "file_type": "creditors",
            "row_count": len(ledger_rows),
            "party_count": len(summary_parties),
            "status": "success",
        }).execute()

        elapsed = time.time() - t_start
        msg = (
            f"✅ Creditors synced — {len(summary_parties)} parties, "
            f"{len(ledger_rows)} ledger rows, "
            f"{len(outstanding_rows)} outstanding rows "
            f"({elapsed:.1f}s)"
        )
        _log(log_callback, msg)
        return True, msg

    except Exception as exc:
        msg = f"❌ Creditors sync failed: {exc}"
        _log(log_callback, msg)
        try:
            _admin_client(supabase_url, service_role_key).table("fin_sync_log").insert({
                "file_type": "creditors",
                "row_count": 0,
                "party_count": 0,
                "status": f"error: {exc}",
            }).execute()
        except Exception:
            pass
        return False, msg


def sync_address_book(
    excel_path: str,
    supabase_url: str,
    service_role_key: str,
    log_callback=None,
) -> tuple[bool, str]:
    """
    Parse Address Book Excel and upsert to Supabase fin_address.
    Uses UPSERT on party_name (not delete-all) to preserve data for groups
    beyond debtors/creditors.
    Returns (success: bool, summary_message: str).
    """
    t_start = time.time()

    try:
        _log(log_callback, "📂 Loading Address Book Excel…")
        wb = load_workbook(excel_path, read_only=True, data_only=True)

        # The sheet may be named 'Address Book' or similar
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]

        _log(log_callback, f"📊 Parsing Address Book (sheet: '{sheet_name}')…")

        rows_out = []
        # Row 1: title row (skip). Row 2: headers. Row 3+: data.
        for raw_row in ws.iter_rows(min_row=3, values_only=True):
            row = tuple(raw_row) + (None,) * 14
            row = row[:14]

            # Sr No is col 0, Party Name is col 1
            party_name = str(row[1]).strip() if row[1] is not None else ""
            if not party_name:
                continue

            def _str(v):
                if v is None:
                    return None
                s = str(v).strip()
                return s if s else None

            rows_out.append({
                "party_name": party_name,
                "address": _str(row[2]),
                "party_group": _str(row[3]),
                "pincode": _str(row[4]),
                "state_name": _str(row[5]),
                "contact_person": _str(row[6]),
                "phone": _str(row[7]),
                "mobile": _str(row[8]),
                "email": _str(row[9]),
                "website": _str(row[10]),
                "pan_no": _str(row[11]),
                "gstin": _str(row[12]),
                "reg_type": _str(row[13]),
            })

        wb.close()

        _log(log_callback, f"  → Found {len(rows_out)} address book entries.")

        _log(log_callback, "🔌 Connecting to Supabase…")
        sb = _admin_client(supabase_url, service_role_key)

        _log(log_callback, f"📤 Upserting {len(rows_out)} address entries…")
        total = 0
        for batch in _chunk(rows_out, CHUNK_SIZE):
            sb.table("fin_address").upsert(batch, on_conflict="party_name").execute()
            total += len(batch)
            _log(log_callback, f"  ↳ Upserted {total}/{len(rows_out)} entries…")

        sb.table("fin_sync_log").insert({
            "file_type": "address_book",
            "row_count": len(rows_out),
            "party_count": len(rows_out),
            "status": "success",
        }).execute()

        elapsed = time.time() - t_start
        msg = f"✅ Address Book synced — {len(rows_out)} entries ({elapsed:.1f}s)"
        _log(log_callback, msg)
        return True, msg

    except Exception as exc:
        msg = f"❌ Address Book sync failed: {exc}"
        _log(log_callback, msg)
        try:
            _admin_client(supabase_url, service_role_key).table("fin_sync_log").insert({
                "file_type": "address_book",
                "row_count": 0,
                "party_count": 0,
                "status": f"error: {exc}",
            }).execute()
        except Exception:
            pass
        return False, msg


# ── Preview helpers (called before sync to show party/row counts) ──────────────

def preview_debtors(excel_path: str) -> tuple[int, int] | None:
    """
    Quick parse of Summary sheet only to get party + row count for preview.
    Returns (party_count, ledger_row_estimate) or None on error.
    """
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        parties = _parse_summary_sheet(wb["Summary"], "debtor")
        # Count ledger rows without building the full list
        count = sum(
            1
            for row in wb["Ledger with Running Balance"].iter_rows(
                min_row=4, values_only=True
            )
            if not _is_party_name_row(row) and not _is_all_none(row)
        )
        wb.close()
        return len(parties), count
    except Exception:
        return None


def preview_creditors(excel_path: str) -> tuple[int, int] | None:
    """Same as preview_debtors but for creditors."""
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        parties = _parse_summary_sheet(wb["Summary"], "creditor")
        count = sum(
            1
            for row in wb["Ledger with Running Balance"].iter_rows(
                min_row=4, values_only=True
            )
            if not _is_party_name_row(row) and not _is_all_none(row)
        )
        wb.close()
        return len(parties), count
    except Exception:
        return None


def preview_address_book(excel_path: str) -> int | None:
    """Returns entry count or None on error."""
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        count = sum(
            1
            for row in ws.iter_rows(min_row=3, values_only=True)
            if row[1] is not None and str(row[1]).strip()
        )
        wb.close()
        return count
    except Exception:
        return None
