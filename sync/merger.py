"""
merger.py — Incremental Period Merge for PriceMaster Sync.

Merges a short Tally period export (1-3 days) into an existing master
analysed .xlsx, re-runs the analysis pipeline, and returns the merged
vendor_data dict plus merge statistics.

Does NOT write to disk or push to Supabase — the caller handles that.

Usage:
    from merger import merge_period_into_master, MergeStats
    vendor_data, stats = merge_period_into_master(
        master_path="debtors_master.xlsx",
        period_path="sundry_debtors_-_today.xls",
        ledger_type="debtor",
        log_callback=self._log,
    )
"""

import dataclasses
import os
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook

import fin_sync_core
from parser import parse_workbook
from processor import compute_running_balances, run_fifo_matching, detect_anomalies


# ─────────────────────────────────────────────────────────────────────────────
# MergeStats dataclass
# ─────────────────────────────────────────────────────────────────────────────

@dataclasses.dataclass
class MergeStats:
    """Counters and detail lists produced by merge_period_into_master()."""
    parties_merged: int = 0       # parties that had ≥1 new rows added
    parties_skipped: int = 0      # parties with opening-balance mismatch (still merged, not skipped)
    parties_new: int = 0          # brand new parties not in master
    rows_added: int = 0           # total new transaction rows appended
    rows_skipped: int = 0         # duplicate rows skipped (already in master)
    rows_amended: int = 0         # rows replaced due to amount change

    # Detail lists for log reporting
    skipped_parties: list = dataclasses.field(default_factory=list)
    # Each entry: (party_name, master_closing, period_opening)

    new_parties: list = dataclasses.field(default_factory=list)
    # Each entry: party_name

    amended_rows: list = dataclasses.field(default_factory=list)
    # Each entry: (party_name, vch_no, old_amt, new_amt)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _log(callback, msg: str):
    if callback:
        callback(msg)


def _txn_key(txn: tuple) -> tuple:
    """
    Deduplication key: (date_normalised, vch_type_lower, vch_no_lower).
    Only vch_no is used for amendment detection (same vch_no, different amount).
    """
    txn_date, vch_type, vch_no, _narration, _debit, _credit = txn
    d = txn_date.date() if isinstance(txn_date, datetime) else txn_date
    vt = (vch_type or "").strip().lower()
    vn = (vch_no or "").strip().lower()
    return (d, vt, vn)


def _txn_amount(txn: tuple) -> float:
    """Return the non-zero amount (debit or credit) for amendment comparison."""
    return (txn[4] or 0.0) + (txn[5] or 0.0)


def _load_master_vendor_data(master_path: str, ledger_type: str, log_callback=None) -> dict:
    """
    Reconstruct vendor_data from an analysed .xlsx master copy.

    Reads:
      - 'Summary' sheet  → opening_bal, stated_closing per party
      - 'Ledger with Running Balance' sheet → transaction rows per party

    Returns vendor_data dict compatible with processor.py's expectations:
    {
        party_name: {
            "opening": float,
            "stated_closing": float,
            "transactions": [(datetime, vch_type, vch_no, narration, debit, credit), ...]
        }
    }
    """
    wb = load_workbook(master_path, read_only=True, data_only=True)

    required_sheets = ("Summary", "Ledger with Running Balance")
    for sheet in required_sheets:
        if sheet not in wb.sheetnames:
            raise ValueError(
                f"Master file is missing sheet '{sheet}'. "
                f"Was it created by 'Analyse Only'? (Found: {wb.sheetnames})"
            )

    # ── Step 1: read Summary sheet for opening / closing ─────────────────────
    summary_rows = fin_sync_core._parse_summary_sheet(wb["Summary"], ledger_type)
    _log(log_callback, f"→ Loading master ({len(summary_rows)} parties)…")

    vendor_data: dict = {}
    for s in summary_rows:
        vendor_data[s["party_name"]] = {
            "opening": s["opening_bal"],
            "stated_closing": s["closing_bal"],
            "transactions": [],
        }

    # ── Step 2: read Ledger sheet for transactions ────────────────────────────
    ledger_rows = fin_sync_core._parse_ledger_sheet(
        wb["Ledger with Running Balance"],
        ledger_type,
        log_callback=log_callback,
    )

    for row in ledger_rows:
        pname = row["party_name"]

        # If a party appears in the ledger but not in summary (shouldn't happen
        # with well-formed files, but guard it), add a placeholder.
        if pname not in vendor_data:
            vendor_data[pname] = {
                "opening": 0.0,
                "stated_closing": 0.0,
                "transactions": [],
            }

        # Convert txn_date string back to datetime for pipeline compatibility
        txn_date_raw = row["txn_date"]
        if txn_date_raw:
            try:
                txn_date = datetime.strptime(txn_date_raw, "%Y-%m-%d")
            except ValueError:
                txn_date = None
        else:
            txn_date = None

        # Skip Opening/Closing Balance rows that writer.py wrote to the ledger
        vch_type_str = (row["vch_type"] or "").strip()
        if vch_type_str in ("Opening Balance", "Closing Balance"):
            continue

        vendor_data[pname]["transactions"].append((
            txn_date,
            row["vch_type"],
            row["vch_no"],
            row["narration"],
            row["debit"],
            row["credit"],
        ))

    wb.close()
    return vendor_data


# ─────────────────────────────────────────────────────────────────────────────
# Main merge function
# ─────────────────────────────────────────────────────────────────────────────

def merge_period_into_master(
    master_path: str,
    period_path: str,
    ledger_type: str,
    log_callback=None,
) -> tuple:
    """
    Load master vendor_data from the analysed .xlsx master copy,
    parse the period file, merge new transactions, re-run analysis pipeline.

    Parameters
    ----------
    master_path  : str   Path to existing analysed .xlsx (master copy).
    period_path  : str   Path to raw Tally period export (.xls/.xlsx).
    ledger_type  : str   "debtor" or "creditor".
    log_callback : callable | None  Receives log lines (thread-safe expected by caller).

    Returns
    -------
    (merged_vendor_data: dict, stats: MergeStats)

    Does NOT write the file or push to Supabase — caller handles that.
    """
    stats = MergeStats()

    # ── 1. Load master ────────────────────────────────────────────────────────
    _log(log_callback, f"📁 Master: {os.path.basename(master_path)}")
    _log(log_callback, f"📁 Period: {os.path.basename(period_path)}")

    vendor_data = _load_master_vendor_data(master_path, ledger_type, log_callback)

    # ── 2. Parse period file ──────────────────────────────────────────────────
    _log(log_callback, "→ Parsing period file…")
    period_vendors, detected_type = parse_workbook(period_path)
    _log(log_callback, f"  ✓ {len(period_vendors)} parties in period file")

    # ── 3. Compute running balances on master so we have computed_closing ─────
    #    (do it quietly here; we will re-run properly after merging)
    vendor_data = compute_running_balances(vendor_data, ledger_type)

    # ── 4. Merge each period party ────────────────────────────────────────────
    affected_parties: set = set()

    for party_name, period_data in period_vendors.items():

        # Extract only real transactions from period (skip Opening/Closing rows)
        period_txns = [
            t for t in period_data.get("transactions", [])
            if (t[1] or "").strip() not in ("Opening Balance", "Closing Balance")
        ]

        period_opening = period_data.get("opening", 0.0)

        if party_name in vendor_data:
            # ── Existing party ────────────────────────────────────────────────
            master_data = vendor_data[party_name]
            master_closing = master_data.get("computed_closing", 0.0)

            # Cross-check: period opening vs master closing.
            # If they differ by ≥ ₹0.50 it means Tally's period export has a
            # different opening than where our master left off (backdated entries,
            # cancellations, rounding, etc.).  We do NOT skip — we trust the master
            # closing as the correct anchor and simply append new transactions on
            # top of it.  The re-analysis pipeline will recompute everything from
            # the master's transaction history, so the running balances stay correct.
            diff = master_closing - period_opening
            if abs(diff) >= 0.50:
                _log(
                    log_callback,
                    f"  ⚠ Opening mismatch (proceeding): {party_name} — "
                    f"master closing ₹{master_closing:,.2f} / "
                    f"period opening ₹{period_opening:,.2f} "
                    f"(diff ₹{diff:,.2f}) — using master closing as anchor",
                )
                stats.parties_skipped += 1
                stats.skipped_parties.append((party_name, master_closing, period_opening))

            # Build lookup of existing transactions keyed by (date, vch_type, vch_no)
            existing_keys: dict = {}   # key → index in master transactions list
            for idx, txn in enumerate(master_data["transactions"]):
                k = _txn_key(txn)
                existing_keys[k] = idx

            # Build vch_no → index lookup for amendment detection
            existing_vch_no: dict = {}
            for idx, txn in enumerate(master_data["transactions"]):
                vn = (txn[2] or "").strip().lower()
                if vn:
                    existing_vch_no[vn] = idx

            rows_added_this_party = 0

            for new_txn in period_txns:
                key = _txn_key(new_txn)
                vn_lower = (new_txn[2] or "").strip().lower()

                if key in existing_keys:
                    # Exact match on (date, vch_type, vch_no) → duplicate, skip
                    stats.rows_skipped += 1
                    continue

                if vn_lower and vn_lower in existing_vch_no:
                    # Same vch_no exists but key differs → check for amount change
                    old_idx = existing_vch_no[vn_lower]
                    old_txn = master_data["transactions"][old_idx]
                    old_amt = _txn_amount(old_txn)
                    new_amt = _txn_amount(new_txn)

                    if abs(old_amt - new_amt) >= 0.01:
                        # Amended voucher → replace
                        master_data["transactions"][old_idx] = new_txn
                        stats.rows_amended += 1
                        stats.amended_rows.append((party_name, new_txn[2], old_amt, new_amt))
                        _log(
                            log_callback,
                            f"  ✏ Amended: {party_name} — vch_no {new_txn[2]} "
                            f"amount changed ₹{old_amt:,.0f} → ₹{new_amt:,.0f}",
                        )
                        affected_parties.add(party_name)
                        continue
                    else:
                        # Same vch_no, same amount, different date → treat as new row
                        pass

                # New transaction — append
                master_data["transactions"].append(new_txn)
                rows_added_this_party += 1
                stats.rows_added += 1
                affected_parties.add(party_name)

            if rows_added_this_party > 0:
                stats.parties_merged += 1
                _log(
                    log_callback,
                    f"  ✓ Merging {party_name} — {rows_added_this_party} new row(s) added",
                )

            # Update stated_closing after merge
            master_data["stated_closing"] = period_data.get("stated_closing", master_closing)

        else:
            # ── New party ─────────────────────────────────────────────────────
            _log(log_callback, f"  ✨ New party: {party_name}")
            vendor_data[party_name] = {
                "opening": period_opening,
                "stated_closing": period_data.get("stated_closing", 0.0),
                "transactions": list(period_txns),
            }
            stats.parties_new += 1
            stats.new_parties.append(party_name)
            affected_parties.add(party_name)

    # ── 5. Sort all affected party transactions by date ───────────────────────
    for pname in affected_parties:
        vendor_data[pname]["transactions"].sort(
            key=lambda t: (t[0] if t[0] is not None else datetime.min)
        )

    # ── 6. Summary line ───────────────────────────────────────────────────────
    _log(
        log_callback,
        f"→ Merge complete: {stats.parties_merged} merged, {stats.parties_new} new, "
        f"{stats.parties_skipped} opening-mismatch (proceeded anyway), "
        f"{stats.rows_added} rows added, "
        f"{stats.rows_skipped} duplicates, {stats.rows_amended} amended",
    )

    # ── 7. Re-run full analysis pipeline on ALL parties ───────────────────────
    _log(log_callback, "→ Re-running analysis pipeline…")

    vendor_data = compute_running_balances(vendor_data, ledger_type)
    _log(log_callback, "  ✓ Running balances recomputed")

    vendor_data = run_fifo_matching(vendor_data, ledger_type)
    _log(log_callback, "  ✓ FIFO matching done")

    anomalies = detect_anomalies(vendor_data, ledger_type)
    errors   = [a for a in anomalies if a["severity"] == "error"]
    warnings = [a for a in anomalies if a["severity"] == "warning"]
    _log(log_callback, f"  Anomalies: {len(errors)} errors, {len(warnings)} warnings")

    # ── 8. Update stated_closing → computed_closing for all affected parties ──
    for pname in affected_parties:
        if "computed_closing" in vendor_data[pname]:
            vendor_data[pname]["stated_closing"] = vendor_data[pname]["computed_closing"]

    return vendor_data, stats
